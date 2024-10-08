

from buspack.resources.db.DB import DB
from buspack.resources.repository.DBRepository import DBRepository
from buspack.service.maps.mapperDB.MapperDB import MapperDB
import openpyxl
import base64
import tempfile
from openpyxl.worksheet.datavalidation import DataValidation
import openpyxl
import os

# Conectar a la Base de Datos

class UpdateExcel():

    def run():

        db = DB()
        conexion = db.connect()

        # Mapeo objetos y accedo a la BD

        bytesExcel = DBRepository().getExcel(conexion)

        if bytesExcel is not None:

            # Obtengo Bytes String en Base64 Del Excel.
            string_base64 = bytesExcel[1]

            # Decodifico el string en base64 a bytes
            bytes_datos = base64.b64decode(string_base64)

            # Guardo los bytes en un archivo temporal
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
                temp_file.write(bytes_datos)
                temp_file_path = temp_file.name

            # Abrir el archivo Excel con openpyxl
            workbook = openpyxl.load_workbook(temp_file_path)
            ws = workbook.active

            # El Excel recuperado de la BD contiene todas su propiedades menos las Validaciones de Datos que no son soportadas por ninguna Biblioteca Actual, 
            # por ende hay que volver a desarrollarlas.

            # Declaro las hojas que me importan del Excel
            parametros = workbook["Parametros"]
            cpa = workbook["CPA"]
            planilla = workbook["Planilla"]

           
            # Elimina todas las filas anteriores en la hoja "Parametros"
                        
            parametros.delete_rows(2, parametros.max_row - 1)

            workbook.save(temp_file_path)


            # Obtengo las localidades a actualizar en la planillas en la hoja Parametros
            localities = MapperDB.mapLocalities(DBRepository.getLocalities(conexion))

            print("Localidades totales", )

            # Me quedo con las activas   
            localitiesActives = list( filter( lambda x : x.isActive == 1 ,localities) )

            fila_inicio = 2  # La primera fila donde se insertarán los datos

            for locality in localitiesActives:

                parametros.cell(row = fila_inicio, column = 1 ).value = locality.enabled_place 
                parametros.cell(row = fila_inicio, column = 2 ).value = locality.locality_name
                parametros.cell(row = fila_inicio, column = 3 ).value = locality.province_name
                parametros.cell(row = fila_inicio, column = 4 ).value = locality.zip_code 

                fila_inicio += 1

            # Obtener los valores de la columna B de la hoja "Parametros"
            #valores_parametros = [cell.value for cell in parametros["B"] if cell.value is not None]
            # Obtengo la cantidad
            #lenparametros = len(valores_parametros)
            #print(lenparametros)
            # Declaro primer Validacion de datos
            #dataValidation = DataValidation(type="list", formula1 = "=Parametros!$A$2:$A$" + str(lenparametros + 1) )

            # Declaro primer Validacion de datos
            dataValidation = DataValidation(type="list", formula1 = "=Parametros!$A$2:$A$" + str(2500) )
            # solo funciona una vez cargado los bytes de una despues no la modifica mas el largo de la lista, entonces el dia que 2000 se quede corta
            # hay que volver a cargar el excel en bytes y aumentar ese largo y la primera vez q corra ya lo actualiza a esa validacion

            # Agrego la Validacion de datos al WorckBook Activo
            ws.add_data_validation(dataValidation)

            # Aplico la Validacion a todas las celdas de la Columna G, desde la 2 celda hasta la 100
            for row in range(1, min(101, ws.max_row + 1)):
                dataValidation.add(ws[f'G{row}'])

            # Declaro segunda Validacion de datos
            dataValidationBool = DataValidation(type="list", formula1 = '"NO,SI"')

            # Agrego la Validacion de datos al WorckBook Activo
            ws.add_data_validation(dataValidationBool)

            # Aplico la Validacion a todas las celdas de la Columna K, desde la 2 celda hasta la 100
            for row in range(1, min(101, ws.max_row + 1)):
                dataValidationBool.add(ws[f'K{row}'])

            # Declaro tercera y ultima Validacion de datos para CPA

            valores_cpa = [cell.value for cell in cpa["B"] if cell.value is not None]

            len_cpa = len(valores_cpa)

            # Declaro primer Validacion de datos
            dataValidationCPA = DataValidation(type="list", formula1 = "=CPA!$B$2:$B$" + str(len_cpa))

            # Agrego la Validacion de datos al WorckBook Activo
            ws.add_data_validation(dataValidationCPA)

            # Aplico la Validacion a todas las celdas de la Columna Q, desde la 2 celda hasta la 100
            for row in range(1, min(101, ws.max_row + 1)):
                dataValidationCPA.add(ws[f'Q{row}'])

            workbook.save(temp_file_path)

            # Sirve para PRUEBAS para poder ver como se va a persistir el Excel en la Base de Datos.
            # Ruta del archivo Excel original


            #ruta_original = "C:/Users/...../...../buspack-updater-backend/buspackProcessBackendAPI/src/excel/planilla.xlsx"

            # Perissito en el archivo Temporal todos los cambios
            workbook.save(temp_file_path)

            # Accedoo al archivo temporal para obetener los bytes del mismo
            with open(temp_file_path, 'rb') as file:
                excel_bytes_modified = file.read()

            # Realizo una Codificacion de este en Base64
            excel_base64_modified = base64.b64encode(excel_bytes_modified).decode('utf-8') 

            # Actualizo el Excel Mofificado en la Base de Datos
            DBRepository.updateExcel(conexion,excel_base64_modified)

            # Cierro el WorckBook
            workbook.close()

            # Elmino el archivo Temporal
            os.remove(temp_file_path)

            db.closeConnection()

            
        else:
            print("No se encontraron datos en la base de datos.")